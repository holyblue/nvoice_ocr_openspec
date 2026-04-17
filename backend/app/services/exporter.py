"""
Export service: bundle/invoice data to Excel and CSV.
"""

import csv
import io
from decimal import Decimal
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from app.models.invoice import Invoice
from app.models.bundle import Bundle
from app.models.category import Category


INVOICE_COLUMNS = [
    ("invoice_number", "發票號碼"),
    ("purchase_date", "消費日期"),
    ("seller_name", "賣方名稱"),
    ("seller_tax_id", "賣方統編"),
    ("amount_untaxed", "未稅額"),
    ("amount_tax", "稅額"),
    ("amount_total", "總金額"),
    ("category_code", "費用分類代碼"),
    ("category_name", "費用分類名稱"),
    ("status", "狀態"),
]


def _invoice_row(invoice: Invoice) -> dict:
    category_code = None
    category_name = None
    if invoice.category:
        category_code = invoice.category.code
        category_name = invoice.category.name

    return {
        "invoice_number": invoice.invoice_number or "",
        "purchase_date": invoice.purchase_date.strftime("%Y-%m-%d") if invoice.purchase_date else "",
        "seller_name": invoice.seller_name or "",
        "seller_tax_id": invoice.seller_tax_id or "",
        "amount_untaxed": float(invoice.amount_untaxed) if invoice.amount_untaxed else 0,
        "amount_tax": float(invoice.amount_tax) if invoice.amount_tax else 0,
        "amount_total": float(invoice.amount_total) if invoice.amount_total else 0,
        "category_code": category_code or "",
        "category_name": category_name or "",
        "status": invoice.status or "",
    }


def _add_invoice_sheet(ws, invoices: list[Invoice], include_totals: bool = True):
    """Populate a worksheet with invoice data."""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    headers = [col[1] for col in INVOICE_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for invoice in invoices:
        row = _invoice_row(invoice)
        ws.append([row[col[0]] for col in INVOICE_COLUMNS])

    if include_totals and invoices:
        total_amount = sum(
            float(inv.amount_total or 0) for inv in invoices
        )
        totals_row = [""] * len(INVOICE_COLUMNS)
        totals_row[0] = "合計"
        total_col_idx = next(i for i, c in enumerate(INVOICE_COLUMNS) if c[0] == "amount_total")
        totals_row[total_col_idx] = total_amount
        ws.append(totals_row)
        # Bold the totals row
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    # Auto-fit columns
    for column in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in column), default=0)
        ws.column_dimensions[column[0].column_letter].width = max(max_len + 2, 12)


def export_bundle_excel(bundle: Bundle, db: Session) -> bytes:
    """Export all invoices in a bundle to Excel bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "發票清單"
    _add_invoice_sheet(ws, bundle.invoices)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_bundle_csv(bundle: Bundle, db: Session) -> bytes:
    """Export all invoices in a bundle to UTF-8 with BOM CSV bytes."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([col[1] for col in INVOICE_COLUMNS])

    for invoice in bundle.invoices:
        row = _invoice_row(invoice)
        writer.writerow([row[col[0]] for col in INVOICE_COLUMNS])

    # Add totals row
    total = sum(float(inv.amount_total or 0) for inv in bundle.invoices)
    totals = [""] * len(INVOICE_COLUMNS)
    totals[0] = "合計"
    total_col_idx = next(i for i, c in enumerate(INVOICE_COLUMNS) if c[0] == "amount_total")
    totals[total_col_idx] = total
    writer.writerow(totals)

    # UTF-8 with BOM
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def export_invoices_excel(
    invoices: list[Invoice],
    include_items: bool = False,
) -> bytes:
    """Export a list of invoices to Excel, optionally including items sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "發票清單"
    _add_invoice_sheet(ws, invoices)

    if include_items:
        ws2 = wb.create_sheet("品項明細")
        item_headers = ["發票號碼", "品項名稱", "數量", "單價", "金額"]
        ws2.append(item_headers)
        for invoice in invoices:
            for item in invoice.items:
                ws2.append([
                    invoice.invoice_number or "",
                    item.item_name,
                    float(item.quantity) if item.quantity else 1,
                    float(item.unit_price) if item.unit_price else "",
                    float(item.amount) if item.amount else "",
                ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
